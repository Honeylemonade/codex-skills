#!/usr/bin/env python3
"""Add built-in dictionary fields to a vocabulary CSV/XLSX table."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any


BASE_FIELDS = [
    "word",
    "count",
    "phonetic",
    "part_of_speech",
    "meaning_zh",
    "explanation_zh",
    "definition_en",
    "root_affix",
    "memory_note",
    "source_indexes",
    "source_examples",
    "source_songs",
    "songs",
]

DICTIONARY_FIELDS = [
    "phonetic",
    "part_of_speech",
    "meaning_zh",
    "explanation_zh",
    "definition_en",
    "root_affix",
    "memory_note",
]

COMMON_ZH = {
    "come": "来；到来；出现",
    "know": "知道；了解；认识",
    "man": "男人；人类",
    "see": "看见；理解；会见",
    "dead": "死的；失效的；完全地",
    "love": "爱；喜爱；爱情",
    "again": "再次；又一次",
    "enchanter": "施魔法的人；令人着迷的人",
    "friend": "朋友",
    "goodbye": "再见；告别",
    "godspeed": "祝一路顺风；祝成功",
    "retreat": "撤退；退避；静修",
    "look": "看；寻找；样子",
    "miss": "想念；错过；未击中",
    "time": "时间；次数；时机",
    "go": "去；走；变成",
    "got": "得到；拥有",
    "back": "回来；背部；支持",
    "up": "向上；起来；提高",
    "why": "为什么",
    "tell": "告诉；辨别",
    "need": "需要；必要",
    "too": "也；太",
}

GENERIC_ROOT_AFFIX = "基础词或常用词，优先结合歌词例句记忆。"
GENERIC_MEMORY_NOTE = "结合出现频率最高的歌词例句记忆。"

BUILTIN_ENTRIES = {
    "retreat": {
        "phonetic": "/rɪˈtriːt/",
        "part_of_speech": "verb/noun",
        "definition_en": "To move back or withdraw; a quiet place or period away from usual activities.",
        "meaning_zh": "撤退；退避；静修",
        "explanation_zh": "歌词语境中通常表示“退后、撤退、不再继续向前”。",
        "root_affix": "re- = back/again；treat 源自拉丁 tractare/tract，含“拉、处理”的历史关联。可理解为“向后撤回”。",
        "memory_note": "re- 联想 back，“retreat”就是往回退；例句里 man will not retreat = 这个人不会撤退。",
    },
    "come": {
        "phonetic": "/kʌm/",
        "part_of_speech": "verb",
        "definition_en": "To move toward or arrive at a place, state, or time.",
        "meaning_zh": "来；到来；出现",
        "explanation_zh": "歌词语境中常表示某个时刻、召唤或对象“到来”。",
        "root_affix": "基础高频动词，无明显现代英语词缀；重点记搭配 come to / come back / come alive。",
        "memory_note": "把 come 和歌词里的方向感一起记：come to me、come to see、time has come。",
    },
    "know": {
        "phonetic": "/noʊ/",
        "part_of_speech": "verb",
        "definition_en": "To have information, awareness, or understanding.",
        "meaning_zh": "知道；了解；认识",
        "explanation_zh": "歌词里常表示确认、理解或情感上的明白。",
        "root_affix": "基础高频动词，无明显现代英语词缀。",
        "memory_note": "注意 know 的 k 不发音；结合 I know / you know 的语气记。",
    },
    "love": {
        "phonetic": "/lʌv/",
        "part_of_speech": "noun/verb",
        "definition_en": "A strong feeling of affection; to feel deep affection for someone or something.",
        "meaning_zh": "爱；喜爱；爱情",
        "explanation_zh": "歌词语境中多指爱情、喜欢或深情。",
        "root_affix": "基础情感词，无明显现代英语词缀。",
        "memory_note": "优先记歌词搭配：love me、called love、thing called love。",
    },
    "dead": {
        "phonetic": "/ded/",
        "part_of_speech": "adjective",
        "definition_en": "No longer alive; not working; completely or absolutely.",
        "meaning_zh": "死的；失效的；完全地",
        "explanation_zh": "歌词中 I’m a dead man 可理解为“我完了/像死人一样”。",
        "root_affix": "基础形容词，无明显现代英语词缀。",
        "memory_note": "dead man 在歌词里通常不是字面死亡，也可能是夸张的情绪表达。",
    },
    "enchanter": {
        "phonetic": "/ɪnˈtʃɑːntər/",
        "part_of_speech": "noun",
        "definition_en": "A person who enchants, charms, or delights others.",
        "meaning_zh": "施魔法的人；令人着迷的人",
        "explanation_zh": "歌词中像是在呼唤一个带有魔法感或强烈吸引力的人。",
        "root_affix": "enchant = 使着迷、施魔法；-er = 做某事的人，所以 enchanter 可理解为“施魔法/使人着迷的人”。",
        "memory_note": "看到 Enchanter, come to me 可以按“有魔力的人，来到我身边”来记。",
    },
}

ROOT_RULES = [
    (re.compile(r"^re(.+)"), "前缀 re- = again/back，表示“再次、返回、向后”。"),
    (re.compile(r"^un(.+)"), "前缀 un- = not/reverse，表示“不、相反、取消”。"),
    (re.compile(r"^dis(.+)"), "前缀 dis- = apart/not，表示“分开、否定、相反”。"),
    (re.compile(r"^pre(.+)"), "前缀 pre- = before，表示“预先、在前”。"),
    (re.compile(r"^pro(.+)"), "前缀 pro- = forward/for，表示“向前、支持”。"),
    (re.compile(r"(.+)ing$"), "后缀 -ing 表示动作进行、动名词或形容词化。"),
    (re.compile(r"(.+)ed$"), "后缀 -ed 常表示过去、完成或形容词化。"),
    (re.compile(r"(.+)ly$"), "后缀 -ly 常把形容词变为副词，表示“以……方式”。"),
    (re.compile(r"(.+)tion$"), "后缀 -tion 表示名词化，常指动作、状态或结果。"),
    (re.compile(r"(.+)less$"), "后缀 -less = without，表示“没有……的”。"),
    (re.compile(r"(.+)ful$"), "后缀 -ful = full of，表示“充满……的”。"),
    (re.compile(r"(.+)ness$"), "后缀 -ness 表示性质或状态。"),
]


def read_rows(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for xlsx input.") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(value or "").strip() for value in next(rows)]
        return [
            {header[i]: values[i] if i < len(values) and values[i] is not None else "" for i in range(len(header))}
            for values in rows
        ]
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_rows(path: Path, rows: list[dict[str, Any]]) -> None:
    extra_fields = [name for row in rows for name in row.keys() if name not in BASE_FIELDS]
    fieldnames = BASE_FIELDS + sorted(set(extra_fields))
    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for xlsx output.") from exc
        wb = Workbook()
        ws = wb.active
        ws.title = "vocabulary"
        ws.append(fieldnames)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([row.get(name, "") for name in fieldnames])
        widths = {
            "A": 18, "B": 10, "C": 18, "D": 16, "E": 24, "F": 42,
            "G": 52, "H": 42, "I": 42, "J": 28, "K": 72, "L": 42, "M": 48,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"
        wb.save(path)
        return
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def fetch_dictionary(word: str, *, timeout: float) -> dict[str, str]:
    url = f"https://api.dictionaryapi.dev/api/v2/entries/en/{urllib.parse.quote(word)}"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError):
        return {}
    if not isinstance(data, list) or not data:
        return {}
    item = data[0]
    phonetic = item.get("phonetic") or ""
    if not phonetic:
        for phonetic_item in item.get("phonetics") or []:
            phonetic = phonetic_item.get("text") or ""
            if phonetic:
                break
    meanings = item.get("meanings") or []
    part = ""
    definition = ""
    if meanings:
        part = meanings[0].get("partOfSpeech") or ""
        definitions = meanings[0].get("definitions") or []
        if definitions:
            definition = definitions[0].get("definition") or ""
    return {"phonetic": phonetic, "part_of_speech": part, "definition_en": definition}


def root_affix_note(word: str) -> str:
    for pattern, note in ROOT_RULES:
        if pattern.match(word):
            return note
    return GENERIC_ROOT_AFFIX


def memory_note(word: str, definition: str) -> str:
    if word.startswith("re") and len(word) > 4:
        return f"看到 re- 先联想“再次/返回”，再结合例句判断 {word} 的具体动作。"
    if definition:
        return f"把英文释义关键词和歌词例句一起记：{definition[:80]}"
    return GENERIC_MEMORY_NOTE


def enrich_row(row: dict[str, Any], *, fetch_online: bool, cache: dict[str, Any], timeout: float) -> dict[str, Any]:
    word = str(row.get("word", "")).strip().lower()
    enriched = {key: "" if value is None else value for key, value in row.items()}
    for field in BASE_FIELDS:
        enriched.setdefault(field, "")
    info = dict(BUILTIN_ENTRIES.get(word, {}))
    if fetch_online and word and word not in cache:
        cache[word] = fetch_dictionary(word, timeout=timeout)
    if fetch_online and word and cache.get(word):
        info.update(cache[word])
    for key in ("phonetic", "part_of_speech", "definition_en"):
        if not enriched.get(key) and info.get(key):
            enriched[key] = info[key]
    if not enriched.get("meaning_zh"):
        enriched["meaning_zh"] = info.get("meaning_zh") or COMMON_ZH.get(word, "")
    if not enriched.get("explanation_zh"):
        if info.get("explanation_zh"):
            enriched["explanation_zh"] = info["explanation_zh"]
        elif enriched.get("meaning_zh"):
            enriched["explanation_zh"] = f"歌词语境中可理解为：{enriched['meaning_zh']}。"
        else:
            enriched["explanation_zh"] = ""
    if not enriched.get("root_affix") or (info.get("root_affix") and enriched.get("root_affix") == GENERIC_ROOT_AFFIX):
        enriched["root_affix"] = info.get("root_affix") or root_affix_note(word)
    if not enriched.get("memory_note") or (info.get("memory_note") and enriched.get("memory_note") == GENERIC_MEMORY_NOTE):
        enriched["memory_note"] = info.get("memory_note") or memory_note(word, str(enriched.get("definition_en", "")))
    return enriched


def merge_previous_dictionary(rows: list[dict[str, Any]], previous_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    previous_by_word = {str(row.get("word", "")).strip().lower(): row for row in previous_rows}
    merged: list[dict[str, Any]] = []
    for row in rows:
        word = str(row.get("word", "")).strip().lower()
        previous = previous_by_word.get(word, {})
        item = dict(row)
        for field in DICTIONARY_FIELDS:
            if not item.get(field) and previous.get(field):
                item[field] = previous[field]
        merged.append(item)
    return merged


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", help="Input CSV/XLSX vocabulary table")
    parser.add_argument("--output", required=True, help="Output CSV/XLSX path")
    parser.add_argument("--offline", action="store_true", help="Skip public dictionary lookup")
    parser.add_argument("--cache", help="Dictionary lookup cache JSON path")
    parser.add_argument("--previous", help="Previous enriched CSV/XLSX table to reuse dictionary fields from")
    parser.add_argument("--timeout", type=float, default=4.0, help="Dictionary request timeout in seconds")
    parser.add_argument("--sleep", type=float, default=0.05, help="Delay between dictionary requests")
    args = parser.parse_args(argv)

    rows = read_rows(Path(args.input))
    if args.previous:
        rows = merge_previous_dictionary(rows, read_rows(Path(args.previous)))
    cache_path = Path(args.cache) if args.cache else Path(args.output).with_suffix(".dictionary-cache.json")
    if cache_path.exists():
        cache = json.loads(cache_path.read_text(encoding="utf-8"))
    else:
        cache = {}
    output_rows = []
    for row in rows:
        output_rows.append(enrich_row(row, fetch_online=not args.offline, cache=cache, timeout=args.timeout))
        if not args.offline and args.sleep:
            time.sleep(args.sleep)
    write_rows(Path(args.output), output_rows)
    if not args.offline:
        cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"rows={len(output_rows)}")
    print(f"output={args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
