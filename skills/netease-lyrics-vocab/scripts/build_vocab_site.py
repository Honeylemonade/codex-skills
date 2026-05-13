#!/usr/bin/env python3
"""Generate a static vocabulary study website from a NetEase lyric vocabulary table."""

from __future__ import annotations

import argparse
import csv
import html
import json
import sys
from pathlib import Path
from typing import Any


FIELDNAMES = [
    "word",
    "count",
    "meaning_zh",
    "definition_en",
    "source_indexes",
    "source_examples",
    "source_songs",
    "songs",
]


def read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [normalize_row(row) for row in csv.DictReader(f)]


def read_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read xlsx input; export CSV instead.") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(value or "").strip() for value in next(rows)]
    output: list[dict[str, Any]] = []
    for values in rows:
        row = {header[i]: values[i] if i < len(values) else "" for i in range(len(header))}
        output.append(normalize_row(row))
    return output


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = {name: "" if row.get(name) is None else str(row.get(name, "")).strip() for name in FIELDNAMES}
    try:
        normalized["count"] = int(float(normalized["count"] or 0))
    except ValueError:
        normalized["count"] = 0
    normalized["examples"] = [item.strip() for item in normalized["source_examples"].split(" | ") if item.strip()]
    normalized["song_list"] = [item.strip() for item in normalized["songs"].split("; ") if item.strip()]
    normalized["source_song_list"] = [item.strip() for item in normalized["source_songs"].split(" | ") if item.strip()]
    normalized["source_index_list"] = [item.strip() for item in normalized["source_indexes"].split("; ") if item.strip()]
    return normalized


def read_vocab(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return read_xlsx(path)
    return read_csv(path)


def json_for_html(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False).replace("</", "<\\/")


def build_html(
    rows: list[dict[str, Any]],
    *,
    title: str,
    source_name: str,
    audio_manifest: dict[str, Any] | None = None,
) -> str:
    payload = json_for_html(
        {
            "title": title,
            "sourceName": source_name,
            "rows": rows,
            "audio": audio_manifest or {},
        }
    )
    escaped_title = html.escape(title)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escaped_title}</title>
  <style>
    :root {{
      color-scheme: light;
      --bg: #f7f8fa;
      --panel: #ffffff;
      --panel-2: #f0f3f6;
      --text: #17202a;
      --muted: #697586;
      --border: #d8dee6;
      --accent: #1f7a5c;
      --accent-soft: #e4f4ed;
      --danger: #a34539;
      --shadow: 0 8px 24px rgba(20, 32, 45, 0.08);
    }}

    * {{ box-sizing: border-box; }}

    html, body {{
      height: 100%;
    }}

    body {{
      margin: 0;
      min-height: 100%;
      overflow: hidden;
      background: var(--bg);
      color: var(--text);
      font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      letter-spacing: 0;
    }}

    button, input, select {{
      font: inherit;
    }}

    button {{
      border: 1px solid var(--border);
      background: var(--panel);
      color: var(--text);
      min-height: 36px;
      border-radius: 7px;
      padding: 0 12px;
      cursor: pointer;
    }}

    button:hover {{ border-color: var(--accent); }}
    button.primary {{ background: var(--accent); color: white; border-color: var(--accent); }}
    button.icon {{ width: 36px; padding: 0; }}
    button.active {{ background: var(--accent-soft); border-color: var(--accent); color: var(--accent); }}

    .app {{
      display: grid;
      grid-template-columns: minmax(280px, 380px) minmax(0, 1fr);
      height: 100vh;
      min-height: 0;
      overflow: hidden;
    }}

    aside {{
      border-right: 1px solid var(--border);
      background: var(--panel);
      padding: 18px;
      display: flex;
      flex-direction: column;
      gap: 14px;
      min-width: 0;
      min-height: 0;
      overflow: hidden;
    }}

    main {{
      padding: 24px;
      min-width: 0;
      display: grid;
      grid-template-rows: auto minmax(0, 1fr);
      gap: 18px;
      height: 100vh;
      overflow: hidden;
    }}

    h1 {{
      margin: 0;
      font-size: 22px;
      line-height: 1.2;
      font-weight: 760;
    }}

    .meta, .muted {{
      color: var(--muted);
      font-size: 13px;
      line-height: 1.4;
    }}

    .stats {{
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 8px;
    }}

    .stat {{
      background: var(--panel-2);
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 10px;
      min-width: 0;
    }}

    .stat strong {{
      display: block;
      font-size: 18px;
      line-height: 1.1;
    }}

    .controls {{
      display: grid;
      gap: 10px;
    }}

    .control-row {{
      display: grid;
      grid-template-columns: 1fr 96px;
      gap: 8px;
    }}

    input, select {{
      width: 100%;
      border: 1px solid var(--border);
      background: white;
      color: var(--text);
      border-radius: 7px;
      min-height: 38px;
      padding: 0 10px;
    }}

    .list {{
      overflow: auto;
      border: 1px solid var(--border);
      border-radius: 8px;
      min-height: 240px;
      flex: 1 1 auto;
    }}

    .word-row {{
      width: 100%;
      border: 0;
      border-bottom: 1px solid var(--border);
      border-radius: 0;
      padding: 10px 12px;
      min-height: 54px;
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 10px;
      text-align: left;
      background: white;
    }}

    .word-row:last-child {{ border-bottom: 0; }}
    .word-row.active {{ background: var(--accent-soft); }}
    .word-row .word {{ font-weight: 720; overflow-wrap: anywhere; }}
    .word-row .count {{ color: var(--muted); font-variant-numeric: tabular-nums; }}

    .study {{
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 8px;
      box-shadow: var(--shadow);
      min-height: 0;
      height: calc(100vh - 48px);
      overflow: hidden;
      display: grid;
      grid-template-rows: auto auto minmax(0, 1fr);
    }}

    .card-head {{
      padding: 22px;
      border-bottom: 1px solid var(--border);
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 16px;
      align-items: start;
    }}

    .term {{
      font-size: clamp(34px, 6vw, 72px);
      line-height: 1;
      font-weight: 800;
      overflow-wrap: anywhere;
    }}

    .meaning {{
      margin-top: 12px;
      font-size: 20px;
      line-height: 1.35;
      min-height: 28px;
    }}

    .definition {{
      margin-top: 8px;
      color: var(--muted);
      line-height: 1.45;
      max-width: 820px;
    }}

    .actions {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      justify-content: flex-end;
    }}

    .nav {{
      padding: 12px 22px;
      border-bottom: 1px solid var(--border);
      display: flex;
      gap: 8px;
      align-items: center;
      justify-content: space-between;
      flex-wrap: wrap;
    }}

    .examples {{
      overflow: auto;
      padding: 18px 22px 22px;
      display: grid;
      gap: 12px;
      align-content: start;
    }}

    .example {{
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 14px;
      display: grid;
      gap: 10px;
    }}

    .example-line {{
      line-height: 1.55;
      font-size: 16px;
    }}

    mark {{
      background: #ffe9a6;
      color: inherit;
      padding: 0 2px;
      border-radius: 3px;
    }}

    .tagline {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      color: var(--muted);
      font-size: 13px;
    }}

    .tag {{
      background: var(--panel-2);
      border: 1px solid var(--border);
      border-radius: 999px;
      padding: 4px 8px;
    }}

    .empty {{
      padding: 24px;
      color: var(--muted);
      text-align: center;
    }}

    @media (max-width: 860px) {{
      body {{ overflow: auto; }}
      .app {{ grid-template-columns: 1fr; height: auto; min-height: 100vh; overflow: visible; }}
      aside {{ border-right: 0; border-bottom: 1px solid var(--border); max-height: 48vh; }}
      main {{ padding: 14px; height: auto; min-height: 52vh; overflow: visible; }}
      .study {{ height: auto; max-height: none; }}
      .card-head {{ grid-template-columns: 1fr; }}
      .actions {{ justify-content: flex-start; }}
      .stats {{ grid-template-columns: repeat(3, minmax(0, 1fr)); }}
    }}
  </style>
</head>
<body>
  <div class="app">
    <aside>
      <div>
        <h1 id="title"></h1>
        <div class="meta" id="sourceName"></div>
      </div>
      <div class="stats">
        <div class="stat"><strong id="totalWords">0</strong><span class="meta">单词</span></div>
        <div class="stat"><strong id="visibleWords">0</strong><span class="meta">当前</span></div>
        <div class="stat"><strong id="knownWords">0</strong><span class="meta">已会</span></div>
      </div>
      <div class="controls">
        <input id="search" type="search" placeholder="搜索单词、含义、歌曲">
        <div class="control-row">
          <select id="mode">
            <option value="all">全部</option>
            <option value="unknown">未标记</option>
            <option value="known">已会</option>
          </select>
          <input id="minCount" type="number" min="1" value="1" aria-label="最低次数">
        </div>
      </div>
      <div class="list" id="wordList"></div>
    </aside>

    <main>
      <section class="study" id="study">
        <div class="card-head">
          <div>
            <div class="term" id="term"></div>
            <div class="meaning" id="meaning"></div>
            <div class="definition" id="definition"></div>
          </div>
          <div class="actions">
            <button class="icon" id="speakWord" title="朗读单词" aria-label="朗读单词">▶</button>
            <button id="toggleKnown" title="标记已会">✓ 已会</button>
            <button id="hideMeaning" title="显示或隐藏含义">隐藏含义</button>
          </div>
        </div>
        <div class="nav">
          <div class="tagline" id="facts"></div>
          <div class="actions">
            <button id="prevWord">上一词</button>
            <button class="primary" id="nextWord">下一词</button>
          </div>
        </div>
        <div class="examples" id="examples"></div>
      </section>
    </main>
  </div>

  <script id="payload" type="application/json">{payload}</script>
  <script>
    const payload = JSON.parse(document.getElementById('payload').textContent);
    const rows = payload.rows.map((row, index) => ({{ ...row, index }}));
    const audio = payload.audio || {{}};
    const knownKey = `vocab-known:${{payload.sourceName}}:${{rows.length}}`;
    const known = new Set(JSON.parse(localStorage.getItem(knownKey) || '[]'));
    let filtered = [];
    let selected = 0;
    let meaningsVisible = true;
    let voices = [];
    let preferredVoice = null;

    const $ = (id) => document.getElementById(id);
    const collator = new Intl.Collator('en');

    function saveKnown() {{
      localStorage.setItem(knownKey, JSON.stringify([...known]));
    }}

    function chooseVoice() {{
      if (!window.speechSynthesis) return null;
      voices = window.speechSynthesis.getVoices();
      const englishVoices = voices.filter((voice) => /^en[-_]/i.test(voice.lang || ''));
      const preferredNames = [
        'Samantha', 'Ava', 'Allison', 'Victoria', 'Karen',
        'Google US English', 'Google UK English Female',
        'Microsoft Aria', 'Microsoft Jenny', 'Daniel'
      ];
      preferredVoice =
        preferredNames.map((name) => englishVoices.find((voice) => voice.name.includes(name))).find(Boolean) ||
        englishVoices.find((voice) => /natural|premium|enhanced/i.test(voice.name)) ||
        englishVoices.find((voice) => /en[-_]US/i.test(voice.lang)) ||
        englishVoices[0] ||
        null;
      return preferredVoice;
    }}

    function speakBrowser(text, kind = 'sentence') {{
      if (!text || !window.speechSynthesis) return;
      window.speechSynthesis.cancel();
      const utterance = new SpeechSynthesisUtterance(text);
      const voice = preferredVoice || chooseVoice();
      if (voice) {{
        utterance.voice = voice;
        utterance.lang = voice.lang || 'en-US';
      }} else {{
        utterance.lang = 'en-US';
      }}
      utterance.rate = kind === 'word' ? 0.72 : 0.82;
      utterance.pitch = 1.0;
      window.speechSynthesis.speak(utterance);
    }}

    function playAudio(src, fallbackText, kind = 'sentence') {{
      if (!src) {{
        speakBrowser(fallbackText, kind);
        return;
      }}
      if (window.speechSynthesis) window.speechSynthesis.cancel();
      const player = new Audio(src);
      player.play().catch(() => speakBrowser(fallbackText, kind));
    }}

    function highlight(text, word) {{
      const escaped = text.replace(/[&<>"']/g, (ch) => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[ch]));
      const pattern = new RegExp(`\\\\b(${{word.replace(/[.*+?^${{}}()|[\\]\\\\]/g, '\\\\$&')}})\\\\b`, 'ig');
      return escaped.replace(pattern, '<mark>$1</mark>');
    }}

    function applyFilters() {{
      const query = $('search').value.trim().toLowerCase();
      const mode = $('mode').value;
      const minCount = Number($('minCount').value || 1);
      filtered = rows.filter((row) => {{
        if (row.count < minCount) return false;
        if (mode === 'known' && !known.has(row.word)) return false;
        if (mode === 'unknown' && known.has(row.word)) return false;
        if (!query) return true;
        return [row.word, row.meaning_zh, row.definition_en, row.songs]
          .join(' ')
          .toLowerCase()
          .includes(query);
      }}).sort((a, b) => b.count - a.count || collator.compare(a.word, b.word));
      selected = Math.min(selected, Math.max(0, filtered.length - 1));
      renderList();
      renderCard();
    }}

    function renderStats() {{
      $('totalWords').textContent = rows.length;
      $('visibleWords').textContent = filtered.length;
      $('knownWords').textContent = known.size;
    }}

    function renderList() {{
      renderStats();
      if (!filtered.length) {{
        $('wordList').innerHTML = '<div class="empty">没有匹配的单词</div>';
        return;
      }}
      $('wordList').innerHTML = filtered.map((row, i) => `
        <button class="word-row ${{i === selected ? 'active' : ''}}" data-index="${{i}}">
          <span>
            <span class="word">${{row.word}}</span>
            <span class="meta">${{known.has(row.word) ? ' · 已会' : ''}}</span>
          </span>
          <span class="count">${{row.count}}</span>
        </button>
      `).join('');
      document.querySelectorAll('.word-row').forEach((button) => {{
        button.addEventListener('click', () => {{
          selected = Number(button.dataset.index);
          renderList();
          renderCard();
        }});
      }});
    }}

    function renderCard() {{
      if (!filtered.length) {{
        $('term').textContent = '';
        $('meaning').textContent = '';
        $('definition').textContent = '';
        $('facts').innerHTML = '';
        $('examples').innerHTML = '<div class="empty">调整筛选条件</div>';
        return;
      }}
      const row = filtered[selected];
      $('term').textContent = row.word;
      const meaning = row.meaning_zh || '待补充中文释义';
      $('meaning').textContent = meaningsVisible ? meaning : '••••••';
      $('definition').textContent = meaningsVisible ? row.definition_en || '' : '';
      $('toggleKnown').classList.toggle('active', known.has(row.word));
      $('toggleKnown').textContent = known.has(row.word) ? '✓ 已会' : '✓ 标记';
      $('hideMeaning').textContent = meaningsVisible ? '隐藏含义' : '显示含义';
      $('facts').innerHTML = [
        `<span class="tag">次数 ${{row.count}}</span>`,
        `<span class="tag">位置 ${{row.source_indexes || '-'}}</span>`,
        `<span class="tag">${{selected + 1}} / ${{filtered.length}}</span>`
      ].join('');

      const examples = row.examples && row.examples.length ? row.examples : ['暂无例句'];
      $('examples').innerHTML = examples.map((example, i) => `
        <div class="example">
          <div class="example-line">${{highlight(example, row.word)}}</div>
          <div class="tagline">
            <span class="tag">${{row.source_index_list && row.source_index_list[i] ? row.source_index_list[i] : '-'}}</span>
            <span class="tag">${{row.source_song_list && row.source_song_list[i] ? row.source_song_list[i] : row.songs || 'source'}}</span>
            <button class="icon speak-example" data-example="${{i}}" title="朗读例句" aria-label="朗读例句">▶</button>
          </div>
        </div>
      `).join('');
      document.querySelectorAll('.speak-example').forEach((button) => {{
        button.addEventListener('click', () => {{
          const exampleIndex = Number(button.dataset.example);
          const source = ((audio.examples || {{}})[row.word] || [])[exampleIndex];
          playAudio(source, examples[exampleIndex], 'sentence');
        }});
      }});
    }}

    function move(delta) {{
      if (!filtered.length) return;
      selected = (selected + delta + filtered.length) % filtered.length;
      renderList();
      renderCard();
    }}

    $('title').textContent = payload.title;
    $('sourceName').textContent = payload.sourceName;
    $('search').addEventListener('input', applyFilters);
    $('mode').addEventListener('change', applyFilters);
    $('minCount').addEventListener('input', applyFilters);
    $('speakWord').addEventListener('click', () => {{
      const row = filtered[selected];
      if (!row) return;
      playAudio((audio.words || {{}})[row.word], row.word, 'word');
    }});
    $('prevWord').addEventListener('click', () => move(-1));
    $('nextWord').addEventListener('click', () => move(1));
    $('toggleKnown').addEventListener('click', () => {{
      const row = filtered[selected];
      if (!row) return;
      if (known.has(row.word)) known.delete(row.word);
      else known.add(row.word);
      saveKnown();
      applyFilters();
    }});
    $('hideMeaning').addEventListener('click', () => {{
      meaningsVisible = !meaningsVisible;
      renderCard();
    }});
    if (window.speechSynthesis) {{
      chooseVoice();
      window.speechSynthesis.onvoiceschanged = chooseVoice;
    }}
    document.addEventListener('keydown', (event) => {{
      if (event.target.matches('input, select')) return;
      if (event.key === 'ArrowRight') move(1);
      if (event.key === 'ArrowLeft') move(-1);
      if (event.key === ' ') {{
        event.preventDefault();
        const row = filtered[selected];
        row && playAudio((audio.words || {{}})[row.word], row.word, 'word');
      }}
    }});
    applyFilters();
  </script>
</body>
</html>
"""


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", help="Vocabulary CSV/XLSX generated by netease_lyrics_vocab.py")
    parser.add_argument("--output", default="vocab-study-site/index.html", help="Output HTML path")
    parser.add_argument("--title", default="歌词单词学习", help="Page title")
    parser.add_argument("--audio-manifest", help="Optional JSON manifest generated by generate_tts_audio.py")
    args = parser.parse_args(argv)

    input_path = Path(args.input)
    rows = read_vocab(input_path)
    audio_manifest = None
    if args.audio_manifest:
        audio_manifest = json.loads(Path(args.audio_manifest).read_text(encoding="utf-8"))
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        build_html(rows, title=args.title, source_name=input_path.name, audio_manifest=audio_manifest),
        encoding="utf-8",
    )
    print(f"input={input_path}")
    print(f"words={len(rows)}")
    print(f"output={output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
